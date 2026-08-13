# -*- coding: utf-8 -*-
"""
RADAR AMARANTE -- Collecteur DFC (U.S. International Development Finance Corp).
==============================================================================

POURQUOI CETTE SOURCE
---------------------
La DFC finance des projets du SECTEUR PRIVE dans les pays en developpement
(energie, infrastructure, sante, industrie, agro). Chaque projet NOMME un
borrower prive (`Project Name`), avec pays, secteur NAICS, montant engage,
description riche, categorie E&S et un flag Sovereign (Yes/No). Un borrower
prive etranger qui deploie en zone a risque = coeur de cible Amarante, exactement
comme IFC/MIGA/Proparco.

Meme DOCTRINE que Proparco :
  - Filtre PERIMETRE pays (zones a risque couvertes ; noms ANGLAIS -> ISO3).
  - Filtre FI : on ECARTE le NAICS "Finance and Insurance" (intermediation, pas
    de deploiement terrain) -- meme crible que MIGA/IFC/Proparco.
  - Filtre FRAICHEUR : Fiscal Year recent (le fichier contient du legacy des
    annees 1960 -- OPIC -- sans interet operationnel aujourd'hui).
  - Scoring LLM (Haiku + escalade Sonnet) via le coeur TED.

VOIE D'ACCES VERIFIEE (sonde_dfc.py, sortie reelle)
---------------------------------------------------
Excel annuel statique, ouvert, sans auth (maj annuelle ~45 j apres cloture) :
  https://www.dfc.gov/sites/default/files/media/documents/
      FY24%20DFC%20Annual%20Project%20Data_508.xlsx
Feuille 'Project Data', en-tete en ligne 1 (une ligne titre au-dessus). Colonnes
confirmees : Fiscal Year, Project Number, Project Type, Region, Country,
Project Name, Committed, NAICS Sector, Project Description, Project Profile URL,
Sovereign (Yes/No), Environmental and Social Risk Category, ...

DEPENDANCE : openpyxl (a ajouter a requirements-radar.txt).

POSTURE : RADAR_DFC_DEBUG=1 valide l'entonnoir SANS LLM ni ecriture (motif
IsDB/IDB) et REVELE les pays non mappes. Isole : un echec ici n'affecte rien.
    RADAR_DFC_DEBUG=1 python dfc_radar.py
"""

import json
import os
import re
import sys
import time
from datetime import date

import ted_complet_v14 as ted
import radar_resilience

# ===========================================================================
# CONFIGURATION
# ===========================================================================
XLSX = os.environ.get(
    "DFC_XLSX",
    "https://www.dfc.gov/sites/default/files/media/documents/"
    "FY24%20DFC%20Annual%20Project%20Data_508.xlsx")
FEUILLE = "Project Data"

NOM_ONGLET = "dfc_radar"
ACTIVER = os.environ.get("RADAR_DFC", "1") != "0"
DEBUG = os.environ.get("RADAR_DFC_DEBUG", "0") == "1"
MAX_AVIS_LLM = int(os.environ.get("DFC_MAX_LLM", "60"))
FY_MIN = int(os.environ.get("DFC_FY_MIN", "2021"))              # exclut le legacy
TIMEOUT = 120
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36")
ENTETES = {"User-Agent": UA, "Accept": "*/*"}

# Crible FI (secteur NAICS) : pas de deploiement terrain.
MOTS_FI = ("finance", "insurance", "financial", "fund", "banking", "monetary")


def _norm(s):
    return re.sub(r"\s+", " ", str(s or "")).strip().lower()


def est_secteur_financier(naics):
    n = _norm(naics)
    return any(m in n for m in MOTS_FI)


# Projets a nom/secteur masques (DFC redige les projets sensibles, surtout
# Ukraine) : aucune cible actionnable, rien a donner au LLM. Ecartes par defaut,
# reactivables via DFC_GARDER_REDACTED=1.
GARDER_REDACTED = os.environ.get("DFC_GARDER_REDACTED", "0") == "1"


def est_redacted(valeur):
    return "redacted" in _norm(valeur)


# ---------------------------------------------------------------------------
# MAPPING NOM DE PAYS (anglais DFC) -> ISO3. Couvre le perimetre a risque ;
# le DEBUG revele les non-mappes, qu'on ajoute ensuite en alias.
# ---------------------------------------------------------------------------
CARTE_PAYS_EN = {
    # Afrique
    "algeria": "DZA", "angola": "AGO", "benin": "BEN", "botswana": "BWA",
    "burkina faso": "BFA", "burundi": "BDI", "cabo verde": "CPV", "cape verde": "CPV",
    "cameroon": "CMR", "central african republic": "CAF", "chad": "TCD",
    "comoros": "COM", "congo": "COG", "republic of congo": "COG",
    "democratic republic of the congo": "COD", "democratic republic of congo": "COD",
    "congo, democratic republic": "COD", "dr congo": "COD", "drc": "COD",
    "cote d'ivoire": "CIV", "cote d ivoire": "CIV", "ivory coast": "CIV",
    "djibouti": "DJI", "egypt": "EGY", "equatorial guinea": "GNQ",
    "eritrea": "ERI", "eswatini": "SWZ", "swaziland": "SWZ", "ethiopia": "ETH",
    "gabon": "GAB", "gambia": "GMB", "ghana": "GHA", "guinea": "GIN",
    "guinea-bissau": "GNB", "kenya": "KEN", "lesotho": "LSO", "liberia": "LBR",
    "libya": "LBY", "madagascar": "MDG", "malawi": "MWI", "mali": "MLI",
    "mauritania": "MRT", "mauritius": "MUS", "morocco": "MAR", "mozambique": "MOZ",
    "namibia": "NAM", "niger": "NER", "nigeria": "NGA", "rwanda": "RWA",
    "sao tome and principe": "STP", "senegal": "SEN", "seychelles": "SYC",
    "sierra leone": "SLE", "somalia": "SOM", "south africa": "ZAF",
    "south sudan": "SSD", "sudan": "SDN", "tanzania": "TZA", "togo": "TGO",
    "tunisia": "TUN", "uganda": "UGA", "zambia": "ZMB", "zimbabwe": "ZWE",
    # Moyen-Orient
    "bahrain": "BHR", "iran": "IRN", "iraq": "IRQ", "israel": "ISR",
    "jordan": "JOR", "kuwait": "KWT", "lebanon": "LBN", "oman": "OMN",
    "palestine": "PSE", "west bank and gaza": "PSE", "qatar": "QAT",
    "saudi arabia": "SAU", "syria": "SYR", "syrian arab republic": "SYR",
    "turkey": "TUR", "turkiye": "TUR", "united arab emirates": "ARE",
    "yemen": "YEM",
    # Amerique latine
    "argentina": "ARG", "bolivia": "BOL", "brazil": "BRA", "chile": "CHL",
    "colombia": "COL", "ecuador": "ECU", "guyana": "GUY", "paraguay": "PRY",
    "peru": "PER", "suriname": "SUR", "uruguay": "URY", "venezuela": "VEN",
    "honduras": "HND", "guatemala": "GTM", "mexico": "MEX",
    "haiti": "HTI", "jamaica": "JAM", "trinidad and tobago": "TTO",
    "dominican republic": "DOM",
    # Europe de l'Est / Caucase / Asie centrale
    "ukraine": "UKR", "russia": "RUS", "belarus": "BLR", "georgia": "GEO",
    "azerbaijan": "AZE", "kosovo": "XKX", "bosnia and herzegovina": "BIH",
    "serbia": "SRB", "north macedonia": "MKD", "kazakhstan": "KAZ",
    "kyrgyzstan": "KGZ", "kyrgyz republic": "KGZ", "tajikistan": "TJK",
    "turkmenistan": "TKM", "uzbekistan": "UZB", "moldova": "MDA",
    "armenia": "ARM", "albania": "ALB", "montenegro": "MNE",
    # Asie / Pacifique
    "myanmar": "MMR", "burma": "MMR", "sri lanka": "LKA", "nepal": "NPL",
    "philippines": "PHL", "indonesia": "IDN", "cambodia": "KHM", "laos": "LAO",
    "lao pdr": "LAO", "mongolia": "MNG", "pakistan": "PAK", "bangladesh": "BGD",
    "papua new guinea": "PNG", "solomon islands": "SLB", "fiji": "FJI",
    "vanuatu": "VUT",
    # Pays presents chez DFC mais HORS perimetre de risque : mappes pour nettoyer
    # l'alerte "non mappe" ; dans_le_perimetre les exclut de toute facon.
    "india": "IND", "vietnam": "VNM", "viet nam": "VNM", "el salvador": "SLV",
    "belize": "BLZ", "maldives": "MDV", "timor-leste": "TLS", "timor leste": "TLS",
    "greece": "GRC", "bulgaria": "BGR",
}


def iso3_depuis_nom(nom):
    return CARTE_PAYS_EN.get(_norm(nom), "")


# ===========================================================================
# LECTURE DE L'EXCEL (openpyxl ; lignes injectables pour les tests)
# ===========================================================================
def telecharger_lignes(fetch=None):
    """Toutes les lignes de la feuille Project Data. `fetch` injecte (les tests
    passent directement une liste de lignes, sans reseau ni openpyxl)."""
    if fetch is not None:
        return fetch()
    import io
    import openpyxl
    r = ted.session_robuste().get(XLSX, headers=ENTETES, timeout=TIMEOUT)
    r.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    nom = FEUILLE if FEUILLE in wb.sheetnames else wb.sheetnames[-1]
    ws = wb[nom]
    return [["" if c is None else str(c) for c in row]
            for row in ws.iter_rows(values_only=True)]


def _index_entete(lignes):
    """Trouve la ligne d'en-tete (celle qui contient Country ET Project Name)."""
    for i, ligne in enumerate(lignes[:8]):
        bas = [_norm(c) for c in ligne]
        if "country" in bas and any("project name" == x for x in bas):
            return i
    # repli : la ligne la plus large
    return max(range(min(len(lignes), 8)),
               key=lambda i: sum(1 for c in lignes[i] if _norm(c)), default=0)


def _cols(entete):
    """Nom de colonne (normalise) -> index."""
    idx = {}
    for i, c in enumerate(entete):
        cle = _norm(c)
        if cle and cle not in idx:
            idx[cle] = i
    return idx


def _val(ligne, idx, *noms):
    for n in noms:
        i = idx.get(_norm(n))
        if i is not None and i < len(ligne):
            v = ligne[i]
            if v not in (None, ""):
                return str(v).strip()
    return ""


# ===========================================================================
# COLLECTE
# ===========================================================================
def _montant_lisible(brut):
    m = re.sub(r"[^\d.]", "", str(brut or ""))
    try:
        val = float(m)
    except ValueError:
        return "", 0
    return "{:,.0f} USD".format(val).replace(",", " "), int(val)


def rec_vers_avis(ligne, idx):
    pays_en = _val(ligne, idx, "Country")
    iso3 = iso3_depuis_nom(pays_en)
    nom = _val(ligne, idx, "Project Name")
    montant_txt, _m = _montant_lisible(_val(ligne, idx, "Committed", "Exposure"))
    ident = _val(ligne, idx, "Project Number")
    desc = _val(ligne, idx, "Project Description")
    return {
        "publication_number": ident,
        "titre": (nom or "Projet DFC")[:300],
        "acheteur": nom,
        "pays_execution": pays_en,
        "pays_iso3": iso3,
        "secteur": _val(ligne, idx, "NAICS Sector"),
        "categorie_es": _val(ligne, idx, "Environmental and Social Risk Category"),
        "sovereign": _val(ligne, idx, "Sovereign (Yes/No)", "Sovereign"),
        "type_document": "Projet finance (DFC)",
        "statut": "actif (portefeuille)",
        "valeur_estimee": montant_txt,
        "date_publication": _val(ligne, idx, "Fiscal Year"),
        "description": (re.sub(r"\s+", " ", desc)[:ted.MAX_CARACTERES_DESCRIPTION]
                        if desc else ""),
        "lien_avis": _val(ligne, idx, "Project Profile URL"),
    }


def collecte(deja_vus=None, fetch=None):
    """Lit l'Excel, filtre (FY recent + FI + perimetre + memoire), renvoie
    (avis_list, compteurs)."""
    deja_vus = deja_vus or set()
    c = {"lignes": 0, "hors_fy": 0, "rejet_fi": 0, "hors_perimetre": 0,
         "rejet_redacted": 0, "deja_connus": 0, "sans_id": 0, "retenus": 0,
         "pays_non_mappes": {}}
    try:
        lignes = telecharger_lignes(fetch=fetch)
    except Exception as e:
        raise RuntimeError("lecture Excel DFC impossible : {}".format(str(e)[:150]))
    if not lignes:
        return [], c
    i_ent = _index_entete(lignes)
    idx = _cols(lignes[i_ent])
    avis = []
    for ligne in lignes[i_ent + 1:]:
        if not any(_norm(x) for x in ligne):
            continue
        c["lignes"] += 1
        fy = re.sub(r"[^\d]", "", _val(ligne, idx, "Fiscal Year"))
        if fy and int(fy) < FY_MIN:
            c["hors_fy"] += 1
            continue
        if est_secteur_financier(_val(ligne, idx, "NAICS Sector")):
            c["rejet_fi"] += 1
            continue
        pays_en = _val(ligne, idx, "Country")
        iso3 = iso3_depuis_nom(pays_en)
        if not iso3 or not ted.dans_le_perimetre(iso3):
            c["hors_perimetre"] += 1
            if pays_en and not iso3:
                c["pays_non_mappes"][pays_en] = c["pays_non_mappes"].get(pays_en, 0) + 1
            continue
        a = rec_vers_avis(ligne, idx)
        if not GARDER_REDACTED and (est_redacted(a["acheteur"]) or est_redacted(a["secteur"])):
            c["rejet_redacted"] += 1
            continue
        if not a["publication_number"]:
            c["sans_id"] += 1
            continue
        if a["publication_number"] in deja_vus:
            c["deja_connus"] += 1
            continue
        avis.append(a)
    c["retenus"] = len(avis)
    # Plus gros montants d'abord.
    def _m(a):
        x = re.sub(r"[^\d]", "", a.get("valeur_estimee") or "")
        return int(x) if x else 0
    avis.sort(key=_m, reverse=True)
    return avis, c


# ===========================================================================
# SCORING (coeur TED) + PROMPT
# ===========================================================================
def avis_pour_scoring(avis):
    copie = dict(avis)
    copie["pays_execution"] = avis.get("pays_iso3") or avis.get("pays_execution")
    return copie


def cible_commerciale(avis, extraction):
    ent = avis.get("acheteur") or "l'entreprise projet"
    pays = avis.get("pays_execution") or "le pays hote"
    return ("Entreprise privee financee par la DFC (US) : {} en {}. Cible : le "
            "borrower et ses equipes deployees sur zone a risque.").format(ent, pays)


PROMPT_DFC = """Tu es analyste sûreté pour une société française de protection de personnes en zones à risque (escorte, protection rapprochée CPO/CPD, chauffeur sécurité, véhicule sécurisé, sécurisation de déplacements terrain). Elle ne vend PAS de conseil voyage générique.

On te donne un PROJET D'INVESTISSEMENT PRIVÉ financé par la DFC (institution de financement du développement des États-Unis) dans un pays hôte à risque. Le borrower est le plus souvent une entreprise privée qui déploie cadres, techniciens et actifs sur le terrain. Détermine si le projet implique une présence PHYSIQUE de personnel sur zone, créant un besoin probable de prestations opérationnelles de sûreté.

RÈGLE DÉPLOIEMENT : un projet industriel, énergétique, minier, d'infrastructure, agro-industriel ou de construction implique des actifs et des équipes sur site. Un projet purement financier (intermédiation, fonds, assurance) sans chantier n'expose pas de personnel.

RÈGLE MOBILITÉ TERRAIN : classe dans UNE catégorie : aucune | capitale | multi_sites | chantier | terrain_isole | frontiere.

RÈGLE SÉCURITÉ EXISTANTE : "securite_existante" = aucune | interne_client | prestataire_tiers | inconnu. "prestataire_tiers" n'est PAS un motif d'exclusion (opportunité de déplacement concurrentiel).

RÈGLE CLIENT : le borrower est un acteur PRIVÉ, commercialement accessible (sauf mention Sovereign=Yes, qui indique un emprunteur souverain/étatique, marché moins accessible).

RÈGLE PROFILS : ne cite JAMAIS d'entreprise réelle en plus de celle donnée, décris des PROFILS d'acteur.

Réponds UNIQUEMENT en JSON valide, sans texte ni Markdown autour, sans commentaire entre parenthèses dans les valeurs.

Schéma :
{{
  "deploiement_terrain_reel": true | false,
  "type_mobilite": "aucune | capitale | multi_sites | chantier | terrain_isole | frontiere",
  "profil_personnes_exposees": "expert_international | executive | technicien | ouvrier_local | aucun",
  "securite_existante": "aucune | interne_client | prestataire_tiers | inconnu",
  "indices_deploiement": ["courtes citations"],
  "type_activite": "assistance_technique | supervision_chantier | etude_terrain | fourniture_equipement | formation | autre",
  "type_client": "bailleur_donateur | institution_ue_onu | etat_administration_locale | entreprise_privee | autre",
  "duree_estimee": "courte_ponctuelle | longue_ou_residente | indetermine",
  "accessibilite_commerciale": "facile | moyenne | difficile",
  "profils_acteurs_probables": ["types de profils, jamais de noms reels"],
  "besoin_securite_operationnel_probable": true | false,
  "niveau_opportunite_amarante": "fort | moyen | faible",
  "justification": "une à deux phrases, besoin opérationnel concret",
  "confiance": 0.0 à 1.0
}}

Projet à analyser :
Borrower (entreprise) : {acheteur}
Emprunteur souverain (Sovereign) : {sovereign}
Pays hôte (exécution) : {pays_execution}
Secteur (NAICS) : {secteur}
Contexte : {description}
"""


def analyser(avis, modele=None):
    prompt = PROMPT_DFC.format(
        acheteur=avis.get("acheteur", ""),
        sovereign=avis.get("sovereign", "") or "n.c.",
        pays_execution=avis.get("pays_execution", ""),
        secteur=avis.get("secteur", "") or "n.c.",
        description=avis.get("description", "") or "(non fournie)",
    )
    texte = ted.appeler_modele(prompt, modele=modele)
    if texte is None:
        return None
    try:
        return ted.normaliser_securite(json.loads(texte))
    except json.JSONDecodeError:
        pass
    debut, fin = texte.find("{"), texte.rfind("}")
    if debut != -1 and fin != -1 and fin > debut:
        try:
            return ted.normaliser_securite(json.loads(texte[debut:fin + 1]))
        except json.JSONDecodeError:
            pass
    repare = ted.reparer_json(texte, modele=ted.MODELE_RAFFINEMENT)
    if repare is None:
        return None
    try:
        return ted.normaliser_securite(json.loads(repare))
    except json.JSONDecodeError:
        return None


# ===========================================================================
# SORTIE (onglet dfc_radar) + miroir Postgres
# ===========================================================================
COLONNES = [
    "date_maj", "score_final", "score_surete", "score_commercial",
    "action_recommandee", "fenetre_action", "niveau_opportunite_amarante",
    "titre", "acheteur", "pays_execution", "secteur", "categorie_es",
    "sovereign", "date_publication",
    "type_client", "type_mobilite", "profil_personnes_exposees",
    "duree_estimee", "accessibilite_commerciale", "securite_existante_detectee",
    "profils_acteurs_probables", "cible_commerciale_reelle",
    "justification", "confiance", "modele", "raffine", "divergence",
    "valeur_estimee", "publication_number", "lien_avis",
]
COLONNE_STATUT_SUIVI = "statut_suivi"
COLONNE_DATE_DETECTION = "date_detection"
TOUTES_COLONNES = COLONNES + [COLONNE_STATUT_SUIVI, COLONNE_DATE_DETECTION]


def ouvrir_feuille(sheet_id, fichier_cs):
    import gspread
    classeur = radar_resilience.ouvrir_classeur(sheet_id, fichier_cs)
    try:
        return classeur.worksheet(NOM_ONGLET)
    except gspread.WorksheetNotFound:
        f = classeur.add_worksheet(title=NOM_ONGLET, rows=3000, cols=len(TOUTES_COLONNES))
        f.append_row(TOUTES_COLONNES)
        return f


def ligne_depuis_resultat(r):
    avis, e = r["avis"], (r["extraction"] or {})
    modele = ted.MODELE_RAFFINEMENT if r["raffine"] else ted.MODELE
    v = {
        "date_maj": date.today().isoformat(),
        "score_final": r["score"], "score_surete": r["surete"],
        "score_commercial": r["commercial"],
        "action_recommandee": ted.calculer_action_recommandee(
            r["score"], r["extraction"], surete=r["surete"]),
        "fenetre_action": ted.calculer_fenetre_action(avis),
        "niveau_opportunite_amarante": e.get("niveau_opportunite_amarante", ""),
        "titre": avis.get("titre", ""), "acheteur": avis.get("acheteur", ""),
        "pays_execution": avis.get("pays_execution", ""),
        "secteur": avis.get("secteur", ""),
        "categorie_es": avis.get("categorie_es", ""),
        "sovereign": avis.get("sovereign", ""),
        "date_publication": avis.get("date_publication", ""),
        "type_client": e.get("type_client", ""),
        "type_mobilite": e.get("type_mobilite", ""),
        "profil_personnes_exposees": e.get("profil_personnes_exposees", ""),
        "duree_estimee": e.get("duree_estimee", ""),
        "accessibilite_commerciale": e.get("accessibilite_commerciale", ""),
        "securite_existante_detectee": e.get("securite_existante_detectee", ""),
        "profils_acteurs_probables": ", ".join(e.get("profils_acteurs_probables") or []),
        "cible_commerciale_reelle": cible_commerciale(avis, r["extraction"]),
        "justification": e.get("justification", ""),
        "confiance": e.get("confiance", ""),
        "modele": modele, "raffine": r["raffine"], "divergence": r["divergence"],
        "valeur_estimee": avis.get("valeur_estimee", ""),
        "publication_number": avis.get("publication_number", ""),
        "lien_avis": avis.get("lien_avis", ""),
    }
    return [str(v.get(c, "")) for c in COLONNES]


def ecrire_resultats(feuille, resultats):
    index = ted.charger_index_publication(feuille, COLONNES)
    derniere = ted.lettre_colonne(len(COLONNES))
    maj, nouvelles, nb_n, nb_m = [], [], 0, 0
    for r in resultats:
        pub = r["avis"].get("publication_number", "")
        ligne = ligne_depuis_resultat(r)
        if pub and pub in index:
            maj.append({"range": "A{0}:{1}{0}".format(index[pub], derniere), "values": [ligne]})
            nb_m += 1
        else:
            nouvelles.append(ligne + ["nouveau", date.today().isoformat()])
            nb_n += 1
    if maj:
        radar_resilience.avec_retry(lambda: feuille.batch_update(maj), "ecriture batch_update")
    if nouvelles:
        radar_resilience.avec_retry(
            lambda: feuille.append_rows(nouvelles, value_input_option="RAW"), "ecriture append_rows")
    try:
        import radar_stockage
        plates = [dict(zip(COLONNES, ligne_depuis_resultat(r))) for r in resultats]
        print("  (pg) " + radar_stockage.ecrire_miroir(NOM_ONGLET, plates))
    except Exception as e:
        print("  (pg) miroir indisponible ({})".format(e))
    return nb_n, nb_m


# ===========================================================================
# POINT D'ENTREE
# ===========================================================================
def _afficher_entonnoir(c):
    print("\n--- ENTONNOIR DFC (FY >= {}) ---".format(FY_MIN))
    print("  Lignes de donnees      : {}".format(c["lignes"]))
    print("  Hors FY (legacy)       : {}".format(c["hors_fy"]))
    print("  Rejetes -- secteur FI  : {}".format(c["rejet_fi"]))
    print("  Hors perimetre (pays)  : {}".format(c["hors_perimetre"]))
    print("  Rejetes -- Redacted    : {}".format(c["rejet_redacted"]))
    print("  Deja connus (sautes)   : {}".format(c["deja_connus"]))
    print("  Sans identifiant       : {}".format(c["sans_id"]))
    print("  RETENUS                : {}".format(c["retenus"]))
    if c.get("pays_non_mappes"):
        print("  /!\\ pays NON mappes (a ajouter en alias) :")
        for nom, n in sorted(c["pays_non_mappes"].items(), key=lambda x: -x[1])[:15]:
            print("      {:6}x  {}".format(n, nom))


def main():
    if not ACTIVER:
        print("(info) Collecteur DFC desactive (RADAR_DFC=0).")
        return
    if not DEBUG and not os.environ.get("ANTHROPIC_API_KEY"):
        print("ERREUR : ANTHROPIC_API_KEY absente (ou lance RADAR_DFC_DEBUG=1).")
        return

    sheet_id = os.environ.get("TED_SHEET_ID")
    fichier = os.environ.get("GOOGLE_SERVICE_ACCOUNT_FILE")
    deja_vus = set()
    if not DEBUG and sheet_id and fichier:
        deja_vus = ted.numeros_publication_existants(sheet_id, fichier, NOM_ONGLET, COLONNES) or set()

    print("Etape 1/2 -- Collecte DFC (Excel annuel)...")
    try:
        avis, compteurs = collecte(deja_vus=deja_vus)
    except Exception as e:
        print("ERREUR : {}".format(str(e)[:200]))
        print("(info) Les autres collecteurs et le dashboard ne sont pas affectes.")
        return
    _afficher_entonnoir(compteurs)

    if DEBUG:
        print("\n=== MODE DEBUG : aucun LLM, aucune ecriture ===")
        for i, a in enumerate(avis[:30], start=1):
            print("  {:2}. {} | {} ({}) | {} | {} | Sov={}".format(
                i, (a.get("acheteur") or "?")[:30], (a.get("pays_execution") or "?")[:14],
                a.get("pays_iso3") or "?", (a.get("secteur") or "")[:22],
                a.get("valeur_estimee") or "n.c.", a.get("sovereign") or "?"))
        print("\nRetire RADAR_DFC_DEBUG pour lancer l'analyse reelle.")
        return

    if not avis:
        print("Aucun projet DFC nouveau a analyser ce run.")
        return
    if len(avis) > MAX_AVIS_LLM:
        print("    (plafond {} : {} en attente).".format(MAX_AVIS_LLM, len(avis) - MAX_AVIS_LLM))
        avis = avis[:MAX_AVIS_LLM]

    print("\nEtape 2/2 -- Extraction LLM et score ({} projets, {})...\n".format(len(avis), ted.MODELE))
    resultats = []
    for i, a in enumerate(avis, start=1):
        arret = ted.sortie_selon_sante_llm("dfc")
        if arret:
            print("  " + arret)
            break
        print("[{}/{}] {}...".format(i, len(avis), a["titre"][:60]))
        extraction = analyser(a)
        s, c, f = ted.calculer_scores(avis_pour_scoring(a), extraction)
        resultats.append({"avis": a, "extraction": extraction, "final_haiku": f,
                          "surete": s, "commercial": c, "score": f,
                          "raffine": False, "divergence": False})
        time.sleep(0.4)

    def merite_escalade(r):
        if r["extraction"] is None:
            return False
        return (r["final_haiku"] >= 5 or r["extraction"].get("confiance", 1.0) < 0.7
                or ted.escalade_pour_securite(r["extraction"]))

    a_escalader = [r for r in resultats if merite_escalade(r)]
    if a_escalader:
        print("\n{} projet(s) escalade(s) vers {}...\n".format(len(a_escalader), ted.MODELE_RAFFINEMENT))
        for i, r in enumerate(a_escalader, start=1):
            print("[{}/{}] Raffinement : {}...".format(i, len(a_escalader), r["avis"]["titre"][:60]))
            raffinee = analyser(r["avis"], modele=ted.MODELE_RAFFINEMENT)
            if raffinee is not None:
                s, c, f = ted.calculer_scores(avis_pour_scoring(r["avis"]), raffinee)
                r["extraction"], r["surete"], r["commercial"], r["score"] = raffinee, s, c, f
                r["raffine"] = True
                r["divergence"] = abs(f - r["final_haiku"]) >= 2.0
            time.sleep(0.4)

    resultats.sort(key=lambda r: r["score"], reverse=True)

    if sheet_id and fichier:
        print("\nEcriture dans l'onglet '{}' ({} projets)...".format(NOM_ONGLET, len(resultats)))
        try:
            feuille = ouvrir_feuille(sheet_id, fichier)
            nb_n, nb_m = ecrire_resultats(feuille, resultats)
            print("-> {} nouveau(x), {} mis a jour (statut_suivi jamais touche).".format(nb_n, nb_m))
        except Exception as e:
            print("(dfc) ecriture impossible ({}). Le run continue.".format(e))
    else:
        print("\n(Pas de Sheet : {} projets analyses, affichage seul.)".format(len(resultats)))
        for r in resultats[:15]:
            print("  {:4} | {:28} | {} | {}".format(
                r["score"], (r["avis"].get("acheteur") or "?")[:28],
                (r["avis"].get("pays_execution") or "?")[:14], r["avis"].get("valeur_estimee") or ""))


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("Collecteur DFC interrompu : {}".format(e))
    sys.exit(0)
